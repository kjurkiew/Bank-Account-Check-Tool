import requests
from openpyxl import load_workbook, Workbook
from datetime import date
from typing import List


# Function to get a bank account number
def get_bank_number(nip: str, date: date) -> List[str]:
    """
    Get the bank account numbers associated with a given NIP (Tax Identification Number) and date.

    Args:
        nip (str): The NIP for which bank account numbers are to be retrieved.
        date (date): The date for which the NIP information is requested.

    Returns:
        List[str]: A list of bank account numbers associated with the NIP.
    """
    api_url = f"https://wl-api.mf.gov.pl/api/search/nip/{nip}?date={date}"

    try:
        response = requests.get(api_url)
        response.raise_for_status()  # Raise an exception for HTTP errors (4xx and 5xx)
        data = response.json()

        subject = data.get("result", {}).get("subject")

        if subject is None:
            account_numbers = ["00000000000000000000000000"]
        else:
            account_numbers = subject.get("accountNumbers", [])

        return account_numbers
    except requests.exceptions.RequestException as e:
        # Handle any HTTP request errors here
        print(f"Error making the request: {e}")
    except ValueError as e:
        # Handle JSON decoding errors here
        print(f"Error decoding JSON response: {e}")

    return []  # Return an empty list if there was an error


# Function to check whether any of the downloaded account numbers belong to BNP
def is_bnp(
    nip: str,
    account_numbers: List[str],
    bnp_account_numbers: List[str],
    bank_number: str,
) -> List[str]:
    """
    Check if any of the account numbers belong to BNP Paribas.

    Args:
        nip (str): The NIP (Tax Identification Number) associated with the account numbers.
        account_numbers (List[str]): List of bank account numbers associated with the NIP.
        bnp_account_numbers (List[str]): List of BNP Paribas account numbers found so far.
        bank_number (str): The bank number to compare with.

    Returns:
        List[str]: Updated list of BNP Paribas account numbers.
    """
    for account_number in account_numbers:
        if account_number[2:6] == bank_number:
            bnp_account_numbers.append(nip)
    return bnp_account_numbers


def get_maximum_rows(*, sheet_object) -> int:
    """
    Get the maximum number of rows in an Excel sheet with data.

    Args:
        sheet_object: The Excel sheet object.

    Returns:
        int: The maximum number of rows with data.
    """
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


def process_excel_file(filename: str, bank_number: str):
    """
    Process an Excel file to find NIPs associated with BNP Paribas and save the results to a new Excel file.

    Args:
        filename (str): The name of the input Excel file.
        bank_number (str): The bank number to search for (e.g., '2030' for BNP Paribas).
    """
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    first_row = 2
    last_row = get_maximum_rows(sheet_object=sheet)

    nips = []
    for value in sheet.iter_cols(
        min_row=first_row, max_row=last_row, min_col=7, max_col=7, values_only=True
    ):
        nips.extend(map(str, value))

    date_today = date.today()
    lista_bnp = []

    for nip in nips:
        acc = get_bank_number(nip, date_today)
        is_bnp(nip, acc, lista_bnp, bank_number)
    result_list = "Nipy firm posiadających konto w banku to: " + ", ".join(lista_bnp)
    print(result_list)

    book = Workbook()
    sheet_2 = book.active
    r = 1
    for x in lista_bnp:
        sheet_2.cell(row=r, column=1).value = x
        r += 1

    output_filename = f"wyniki_{first_row}_{last_row}.xlsx"
    book.save(output_filename)


if __name__ == "__main__":
    input_excel_filename = "list_test.xlsx"
    bank_number = "2030"  # BNP PARIBAS
    process_excel_file(input_excel_filename, bank_number)
